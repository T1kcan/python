import openpyxl as xl
from openpyxl.chart import BarChart, Reference
def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']
    cell = sheet['A1']
    cell = sheet.cell(1,1)

    print(cell.value)

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 6)
        print(cell.value)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row,6)
        corrected_price_cell.value = corrected_price

    values = Reference(sheet, 
                min_row=2,
                max_row=sheet.max_row,
                min_col=6,
                max_col=6)

    chart = BarChart
    chart.add_data(values)
    sheet.add_chart(chart, 'g2')

    wb.save(filename)

